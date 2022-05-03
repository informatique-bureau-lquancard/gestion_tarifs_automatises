##########
# IMPORTS 
##########

from ast import Str
import csv
import glob
from operator import index
import re
import io
import time

import openpyxl

from unidecode import unidecode
from lxml import etree
import sys
import xml.etree.ElementTree as ET
# coding: utf-8

# Parser flux xml
import feedparser

# Pour l'authentification de google
import gspread
import pandas as pd

# reader = reader.sort_values(by = [0], axis = 0, ascending=True, inplace=True)
# mis dans les flux google mais non dans les flux xml : à mettre

from openpyxl import Workbook

# Nombre aléatoire
import random

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

import peuplement_de_la_base as pb

import requests

tab_colonne = ['vin', 'millesime', 'formatB', 'prix', 'quantite', 'conditionnement', 'commentaires   ']

##########
# /PROPRIETES
##########
# MILIMA
def fct_flux_tree_xml_1(flux, chemin_fichier_sortie, nom_tarif_profil, id_negociant):

    print ("  fct_flux_tree  ")

    print(flux)

    # ??? normalement il y a déjà flux en haut, supprimer et remplacer par flux ? voir par rapport à la base
    type_flux = "flux_xml"

    print("flux : ")
    print(flux)

    tree = etree.parse(flux)

    Vin = etree.Element("Produit")

    with open(chemin_fichier_sortie,'w',newline='', encoding='utf-8') as monFichierSortie:
        writer = csv.writer(monFichierSortie, delimiter=';', quoting=csv.QUOTE_NONE, escapechar=';', quotechar='')

        taille : int = len( tree.xpath("/XMLFILE") )

        print("taille : " + str(taille) )

        taille : int = len( tree.xpath("/XMLFILE/Produit") )

        print("taille : " + str(taille) )

        for Vin in tree.xpath("/XMLFILE/Produit") :

            # couleur
            couleur : str = (Vin[4].text).upper()

            # vin, enlève la date de vin
            vin : str = unidecode((Vin[0].text)[:-5])

            vin = ft.formaterVin("", vin, couleur)

            # millesime
            millesime : str = ft.formaterAnnee( Vin[1].text )

            # foramtB
            formatB : str = ft.formaterFormatBouteilleBis( Vin[9].text )

            # prix
            qCaisse : str = Vin[18].text
            iNb : int = qCaisse.find("*")
            qCaisse = qCaisse[:int(iNb)].replace(" ","")

            fprix : str = Vin[7].text
            fprix = float(fprix) / int(qCaisse)
            prix : str = str(fprix)
            prix = prix.replace(",",".")

            prix_float : float = round(float(prix), 2)
            prix = str(prix_float)

            # quantite
            quantite : str = Vin[17].text

            # conditionnement

            conditionnement : str = ft.formaterConditionnement(formatB, int(quantite), Vin[10].text, vin)
            conditionnement : str = conditionnement.replace("CCO", 'CC')

            # Mise à jour qui ont rendu inutilisable les conditionnements
            # conditionnement = ft.formatterConditionnement(formatB, quantite, conditionnement, vin)

            # commentaires
            commentaires : str = ""

            if( (conditionnement not in ft.tab_conditionnement) and ( conditionnement == "CBO24DE" ) ):
                commentaires = "VERIF CDT - "+conditionnement

            # insertion des lignes dans le fichier
            newRow=[vin,millesime,formatB,prix,quantite,conditionnement,commentaires]

            # insertion des lignes dans la base
            # pb.insertion_tarif(type_flux, nom_tarif_profil,vin,millesime,formatB,prix,quantite,conditionnement,commentaires, id_negociant)

            writer.writerow(newRow)
        return writer

# ANGWIN_HEBDO
def fct_flux_tree_xml_2(flux, chemin_fichier_sortie, nom_tarif_profil, id_negociant):

    print ("  fct_flux_tree  ")

    print(flux)

    feed = feedparser.parse(flux)

    tree = etree.parse(flux)

    # page = requests.get(flux).text.encode('utf-8')

    # tree = etree.fromstring(page)
    # treepath = tree.xpath('.//table[@class="W(100%)"]')

    # xml = etree.tostring(tree)
    # xml = etree.XML(xml)
    
    # Vin = etree.Element("item")

    with open(chemin_fichier_sortie,'w',newline='', encoding='utf-8') as monFichierSortie:
        writer = csv.writer(monFichierSortie, delimiter=';', quoting=csv.QUOTE_NONE, escapechar=';', quotechar='')

        taille : int = len( treepath )

        print("taille : " + str(taille) )

        treepath = tree.xpath('.//table[@class="W(100%)"]/channel/item')
        
        taille : int = len( treepath )

        print("taille : " + str(taille) )
    
        for entry in feed.entries:

            item = entry.description

            prix_ttc = entry.g_price.replace(' EUR','')
            prix_ht = float(prix_ttc)/float(1.2)
            prix = str(prix_ht).replace('.',',')

            quantite = entry.g_sell_on_google_quantity
            couleur = entry.g_color
        
            rec_millesime = re.findall('[0-9][0-9][0-9][0-9]',item)
            millesime = rec_millesime[0] if rec_millesime else 'NV'
            
            vin = unidecode(item.split(millesime)[0]).replace(';','')
            
            rec_formatB = re.findall('[0-9]?[,]?[0-9][eE0-9][cC]?[lL]',item)

            formatB_dict = ['Magnum','Double Magnum', 'Jeroboam', 'Imper']
            formatB = ''

            for f in formatB_dict :

                if f not in item :
                    break

                iFormat = item.find(f)
                formatB = item[int(iFormat):]

            if rec_formatB and formatB == '':
                formatB = rec_formatB[0]
            elif 'Imperiale' in item :
                formatB = 'IM'
            elif formatB == '' :
                formatB = 'BO'

            formatB : str = ft.formaterFormatBouteille(formatB)

            conditionnement : str = ft.conditionnementParDefaut(formatB,int(quantite))
            commentaires : str = 'Verif CDT'

            print(vin)
            print(millesime)
            print(formatB)
            print(prix)
            print(quantite)
            print(conditionnement)
            print(commentaires)
                
            newRow=[vin, millesime, formatB, prix, quantite, conditionnement, commentaires]
            writer.writerow(newRow)
        return writer

#Initialisation et création du fichier de  xlsx
def initialisation_google_sheet(tab_onglets_ok, flux, chemin_fichier_sortie):

    print ("  initialisation_google_sheet  ")
    print(chemin_fichier_sortie)

    # chemin_fichier_sortie_xlsx = chemin_fichier_sortie+'.xlsx'

    # authentification à google API,  compte j.martin.blq@gmail.com
    # fichier credential.json dans le répertoire du script
    gc = gspread.service_account(filename='/var/www/html/projets/bureau-laurent-quancard/gestion-des-offres/gestion_tarifs_automatises/Tarifs/tests_tarifs/credential.json')

    # ouverture du spreadsheet google de Maion B
    sh = gc.open_by_key(flux)

    # Récupération de la liste des onglets
    worksheet_list = sh.worksheets()

    workbook = pd.ExcelWriter(chemin_fichier_sortie+"1.xlsx", engine = 'openpyxl')
    #pd.to_excel(workbook,sheet_name="Sheet1", startrow=1, startcol=1,index=False, header=True)

    sheets = []

    nombre_onglets_entree = len(worksheet_list)

    print("nombre_onglets_entree : "+str(nombre_onglets_entree))

    #A revoir si le tableau tab_onglets n'est pas renseigné
    #if(len(tab_onglets) < 1):
    index_onglet_sortie = 0

    # Parcours des onglet et création de dataframe Pandas par onglets 
    for index_onglet in range(nombre_onglets_entree) :

        active_worksheet = sh.get_worksheet(index_onglet)
        sheet_title = unidecode(str(active_worksheet).split(' ')[1].replace("'",""))
        
        # Test les onglets à récupérer
        if len(tab_onglets_ok) != 0 :
            if sheet_title not in tab_onglets_ok:
                continue

        sheets.append(sheet_title)   

        print('titre : '+sheet_title)
        print('index_onglet_sortie : '+str(index_onglet_sortie))

        sheets[index_onglet_sortie] = pd.DataFrame(active_worksheet.get_all_records())

        sheets[index_onglet_sortie].to_excel(workbook,sheet_title, index = False)

        index_onglet_sortie = index_onglet_sortie + 1

    workbook.save()

def mise_en_forme_csv(tab_designation, chemin_fichier_sortie):

    with open(chemin_fichier_sortie+'2.csv', newline='', encoding="utf-16") as csvfile :
        reader = csv.DictReader(csvfile)
        with open(chemin_fichier_sortie+"3.csv",'w',newline='', encoding="utf-16") as monFichierSortie:
            writer = csv.writer(monFichierSortie, delimiter=';', quoting=csv.QUOTE_NONE, escapechar=';', quotechar='')
            
            print(tab_designation)

            for row in reader :

                newRow = []

                for index_designation in range(len(tab_designation)):

                    #Test sur la valeur du vin
                    if(index_designation == 0):
                        colonne_vin = unidecode(row[tab_designation[index_designation]]).upper().strip().replace(';','').replace(', ','-')

                        newRow.append(colonne_vin)
                        continue

                    #Test sur la valeur du prix
                    if(index_designation == 3):
                        taille_colonne_prix = len(str(row[tab_designation[index_designation]]))

                        if( taille_colonne_prix < 1):
                            newRow.append(0)
                            continue

                    newRow.append(row[tab_designation[index_designation]])

                writer.writerow(newRow)

def recuperation_colonnes(active_sheet):
    tab_colonnes = []

    max_column : int = active_sheet.max_column

    print("max_column : "+str(max_column))

    #Récupération des colonnes dans un tableau
    for index_colonnes in range(max_column):

        #Utilisation du tableau ASCII : 65 = A

        indice_lettre : int = 65 + index_colonnes

        #print("indice_lettre : "+str(indice_lettre))

        lettre_colonne : str = chr(indice_lettre)

        #print("lettre_colonne : "+lettre_colonne)

        # print(active_sheet[lettre_colonne])

        tab_colonnes.append(active_sheet[lettre_colonne])

    return tab_colonnes

# CUVFAU
def fct_flux_google_sheets1(flux, chemin_fichier_sortie, nom_tarif_profil, id_negociant):

    #Initialisation des onglets faisant parti du fichier final, quand vide prendre tout
    tab_onglets_ok = ['RESULTAT']

    #Initialisation et création du fichier de  xlsx
    initialisation_google_sheet(tab_onglets_ok, flux, chemin_fichier_sortie)

    #transforme en csv afin d'utiliser DictReader
    reader = pd.read_excel(chemin_fichier_sortie+"1.xlsx")

    # reader = reader.sort_values(by = [0], axis = 0, ascending=True, inplace=True)

    reader.to_csv(chemin_fichier_sortie+'2.csv', index = False, encoding="utf-16")

    vin_desi = 'Product Info Name'
    millesime_desi = 'Product Info Millesime'
    formatB_desi = 'Product Info Contenance'
    prix_desi = 'Prices Price Value'
    quantite_desi = 'Stock Disponible 1'
    conditionnement_desi = 'Product Info Unit Conditionnement'
    commentaires_desi = 'Stock Stock Physique'

    appellation = 'Product Info Appellation'
    couleur = 'Product Info Couleur'
    regie = 'Product Info Regie'

    tab_designation = [vin_desi, millesime_desi, formatB_desi, prix_desi, quantite_desi, conditionnement_desi, commentaires_desi, appellation, couleur, regie]

    mise_en_forme_csv(tab_designation, chemin_fichier_sortie)

    #transforme en xlsx afin de pouvoir utiliser openpyxl
    new_pd_reader = pd.read_csv(chemin_fichier_sortie+"3.csv", sep = ';', encoding="utf-16")
    new_pd_reader.to_excel(chemin_fichier_sortie+"4.xlsx", index = False, encoding="utf-16")

    # On load le le tarif au format xlsx dans un workbook
    wb = openpyxl.load_workbook(chemin_fichier_sortie+"4.xlsx")

    # On prépare un nouveau workbook
    wb2 = Workbook()
    dest_filename = chemin_fichier_sortie+'.xlsx'
    ws = wb2.active
    ws.title = nom_tarif_profil

    print("Nombre d'onglet : "+str(len(wb.sheetnames)))

    # Lecture du workbook d'origine
    for index_onglets in range(len(wb.sheetnames)) :

        print(index_onglets)

        sheet_list = wb.sheetnames
        
        active_sheet = wb[sheet_list[index_onglets]]

        tab_colonnes = recuperation_colonnes(active_sheet)

        iVin = tab_colonnes[0]
        iMillesime = tab_colonnes[1]
        iFormatB = tab_colonnes[2]
        iPrix = tab_colonnes[3]
        iQte = tab_colonnes[4]
        iCond = tab_colonnes[5]
        iCom = tab_colonnes[6]
        iAppellation = tab_colonnes[7]
        iCouleur = tab_colonnes[8]
        iRegie = tab_colonnes[9]

        row_count = active_sheet.max_row
        # column_count = active_sheet.max_column

        print("nb lignes : "+str(row_count))

        index_sortie : int = 1

        for index_lignes in range(row_count):

                # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
                # if iVin[i].value is None or str(iVin[i].value) == 'Product Info Name' :
                #     continue
                #else:
            exclude_appellation = ['SAUTERNES','BARSAC']

            # appellation
            appellation : str = unidecode(str(iAppellation[index_lignes].value)).upper()
            appellation = ft.formaterAppellation(appellation, exclude_appellation)

            # couleur
            couleur = str(iCouleur[index_lignes].value).upper()

            # vin
            vin = unidecode( str(iVin[index_lignes].value) )

            vin = ft.formaterVin(appellation, vin, couleur)

            # millesime      
            millesime = iMillesime[index_lignes].value

            prix : str = ft.formaterPrix( iPrix[index_lignes].value )

            #Reconnaissance du format de bouteille, on cherche uniquement les caractères numérique
            formatB = str(iFormatB[index_lignes].value)
            formatB = ft.formaterFormatBouteille(formatB)      

            if iQte[index_lignes].value is None :
                quantite = 0
                continue

            quantite = iQte[index_lignes].value

            rec_cond =str(iCond[index_lignes].value).split(' ')[0].replace('0','').replace('CB','CBO').replace('CT','CC')
            rec_cond = re.findall('[C][BC][O]?[0-9]?[0-9]?',rec_cond)

            if rec_cond :
                conditionnement = rec_cond[0]
            else:
                conditionnement = 'UNITE'
            
            if iCom[index_lignes].value is None :
                commentaires = ''
            else:
                commentaires = 'Stock Phy : '+str(iCom[index_lignes].value)

            #tab_cell = vin millesime formatB prix quantite conditionnement commentaires 
            # tab_cell = fichierFonction.tableau_ligne(tab_colonne, ws, index_lignes)

            #print("index_sortie : "+str(index_sortie))

            ft.affectationLignes(index_sortie, ws, vin, millesime, formatB, prix, quantite, conditionnement, commentaires)

            # Saut de la première ligne
            if index_sortie == 1 :
                index_sortie = index_sortie + 1
                continue

            if vin == ws.cell(row = index_sortie-1, column = 1).value and millesime == ws.cell(row = index_sortie-1, column = 2).value and formatB == ws.cell(row = index_sortie-1, column = 3).value :
                 
                while conditionnement == ws.cell(row = index_sortie-1, column = 6).value :

                    conditionnement = random.choice(ft.tab_conditionnement)
                    commentaires = 'ATTN : VERIF CDT - DOUBLON LIGNE PRECEDENTE - ' + commentaires
            
            index_sortie += 1

    wb2.save(dest_filename)

# MAISOB
def fct_flux_google_sheets2(flux, chemin_fichier_sortie, nom_tarif_profil, id_negociant):

    #Initialisation des onglets faisant parti du fichier final, quand vide prendre tout
    tab_onglets_ok = []
    
    #Initialisation et création du fichier de  xlsx
    initialisation_google_sheet(tab_onglets_ok, flux, chemin_fichier_sortie)

    #transforme en csv afin d'utiliser DictReader
    reader = pd.read_excel(chemin_fichier_sortie+"1.xlsx")

    # reader = reader.sort_values(by = [0], axis = 0, ascending=True, inplace=True)

    reader.to_csv(chemin_fichier_sortie+'2.csv', index = False, encoding="utf-16")

    vin_desi = 'Château / Domaine'
    millesime_desi = 'Millésime'
    formatB_desi = 'Format'
    prix_desi = '€/btl.'
    quantite_desi = 'Qté'
    conditionnement_desi = 'Condit.'
    # commentaires_desi = 'Stock Stock Physique'

    appellation = 'Appellation'
    couleur = 'Couleur'
    regie = 'Régie'

    tab_designation = [vin_desi, millesime_desi, formatB_desi, prix_desi, quantite_desi, conditionnement_desi, appellation, couleur, regie]

    mise_en_forme_csv(tab_designation, chemin_fichier_sortie)

    #transforme en xlsx afin de pouvoir utiliser openpyxl
    new_pd_reader = pd.read_csv(chemin_fichier_sortie+"3.csv", sep = ';', encoding="utf-16")
    new_pd_reader.to_excel(chemin_fichier_sortie+"4.xlsx", index = False, encoding="utf-16")

    # On load le le tarif au format xlsx dans un workbook
    wb = openpyxl.load_workbook(chemin_fichier_sortie+"4.xlsx")

    # On prépare un nouveau workbook
    wb2 = Workbook()
    dest_filename = chemin_fichier_sortie+'.xlsx'
    ws = wb2.active
    ws.title = nom_tarif_profil

    print("Nombre d'onglet : "+str(len(wb.sheetnames)))

    # Lecture du workbook d'origine
    for j in range(len(wb.sheetnames)) :

        print(j)
        sheet_list = wb.sheetnames
        active_sheet = wb[sheet_list[j]]

        tab_colonnes = recuperation_colonnes(active_sheet)

        iVin = tab_colonnes[0]
        iProducteur = ''
        iCouleur = ''
        iMillesime = tab_colonnes[1]
        iFormatB = tab_colonnes[2]
        iPrix = tab_colonnes[3]
        iQte = tab_colonnes[4]
        iCond = tab_colonnes[5]
        # iCom = tab_colonnes[6]
        iAppellation = tab_colonnes[6]
        iCouleur = tab_colonnes[7]
        iRegie = tab_colonnes[8]

        row_count = active_sheet.max_row
        column_count = active_sheet.max_column

        index_sortie : int = 1

        for index_ligne in range(row_count):

            exclude_appellation = []

            appellation : str = ft.formaterAppellation( str(iAppellation[index_ligne].value), exclude_appellation)
            couleur : str = str(iCouleur[index_ligne].value)

            vin : str = unidecode(str(iVin[index_ligne].value))
            vin = ft.formaterVin(appellation, vin, couleur)

            print("   ")
            # print("index : " + str(index_sortie))

            iMillesime_bis : str = re.findall('[0-9][0-9][0-9][0-9]', str(iMillesime[index_ligne].value))

            # print("iMillesime : " + str(iMillesime_bis))

            if( not (len(iMillesime_bis) > 0) ):
                millesime = "NV"
            else:
                millesime : str = ft.formaterAnnee( str(iMillesime_bis[0]) )

            # print("millesime : " + str(millesime))
            
            formatB = str(iFormatB[index_ligne].value)
            formatB = ft.formaterFormatBouteille(formatB)

            prix = float(unidecode(str(iPrix[index_ligne].value)).replace(' ','').replace('EUR',''))

            # print("prix : "+str(prix))

            quantite = iQte[index_ligne].value

            Dict_cond = ['UNITE','CC1','CC2','CC3','CC4','CC6','CC12','CC24DE','CBO1','CBO2','CBO3','CBO6','CBO12','CBO24DE','COLLEC'] # pour le random / doublons
            rec_cond = str(iCond[index_ligne].value).replace('UNIT','UNITE').replace('CT','CC').replace('CO','CC')
            
            if 'COLLECTION' in vin :
                conditionnement = 'COLLEC'
            else:
                conditionnement = rec_cond

            # commentaires = str(iRegie[index_ligne].value)+' '+str(iCom[index_ligne].value)

            commentaires = ""

            ft.affectationLignes(index_sortie, ws, vin, str(millesime), formatB, str(prix), str(quantite), conditionnement, commentaires)

            index_sortie += 1    

    # Fonction de suppression des lignes vides d'un workbook
    index_row = []
    for n in range(1, ws.max_row):
        if ws.cell(n, 1).value is None:
            index_row.append(n)

    for row_del in range(len(index_row)):
        ws.delete_rows(idx=index_row[row_del], amount=1)
        index_row = list(map(lambda k: k -1, index_row))

    wb2.save(dest_filename)